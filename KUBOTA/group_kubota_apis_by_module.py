import os
from typing import List, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_XLSX = os.path.join(BASE_DIR, "KUBOTA_API_Endpoints_Detailed.xlsx")
OUTPUT_XLSX = os.path.join(BASE_DIR, "KUBOTA_API_Endpoints_Detailed_GROUPED.xlsx")


def infer_module(endpoint: str | None) -> str:
    """
    Infer module name from API endpoint path using simple heuristics.
    This won't be perfect but will give a useful logical grouping.
    """
    if not endpoint:
        return "General"

    ep = endpoint.lower()

    # Back-end functional areas (guess from URL segments)
    if "/crm" in ep:
        return "CRM"
    if "/masters" in ep or "/master" in ep:
        return "Masters"
    if "/sales" in ep or "/presales" in ep:
        return "Sales / Pre-Sales"
    if "/service" in ep:
        return "Service"
    if "/spares" in ep or "/parts" in ep:
        return "Spares / Parts"
    if "/training" in ep:
        return "Training"
    if "/warranty" in ep:
        return "Warranty"
    if "/auth" in ep or "/login" in ep or "/security" in ep:
        return "Authentication / Security"
    if "/report" in ep or "/analytics" in ep:
        return "Reports / Analytics"
    if "/file" in ep or "/upload" in ep or "/download" in ep:
        return "File / Document Management"

    # Default catch-all
    return "General"


def load_sheet_rows() -> tuple[List[str], List[Dict[str, str]]]:
    wb = load_workbook(SOURCE_XLSX)
    ws = wb.active

    # First row is assumed to be header
    headers: List[str] = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    # Normalize header names slightly
    normalized = [h.strip() for h in headers]

    # Ensure "Module" column exists
    if "Module" not in normalized:
        normalized.insert(0, "Module")  # add at the beginning

    # Index helpers
    col_index_by_name = {name: idx for idx, name in enumerate(headers)}
    endpoint_col_idx = None
    for name, idx in col_index_by_name.items():
        if name.lower() in {"api endpoint", "endpoint", "path", "url"}:
            endpoint_col_idx = idx
            break

    rows: List[Dict[str, str]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_dict: Dict[str, str] = {}
        # Map existing headers
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = r[idx] if idx < len(r) else None
            row_dict[header] = "" if value is None else str(value)

        # Compute Module
        endpoint_value = None
        if endpoint_col_idx is not None and endpoint_col_idx < len(r):
            endpoint_value = r[endpoint_col_idx]
        elif "API Endpoint" in row_dict:
            endpoint_value = row_dict.get("API Endpoint")

        module = infer_module(str(endpoint_value) if endpoint_value is not None else None)
        row_dict["Module"] = module

        rows.append(row_dict)

    return normalized, rows


def write_grouped_workbook(headers: List[str], rows: List[Dict[str, str]]) -> None:
    # Sort by Module then API Endpoint if present
    def sort_key(r: Dict[str, str]):
        module = r.get("Module", "")
        endpoint = r.get("API Endpoint", r.get("Endpoint", ""))
        return (module, endpoint)

    rows_sorted = sorted(rows, key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "API Endpoints (Grouped)"

    # Write header
    header_font = Font(bold=True)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font

    # Write data
    for row in rows_sorted:
        ws.append([row.get(h, "") for h in headers])

    # Auto-size columns
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Freeze header and add filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_XLSX)


def main():
    headers, rows = load_sheet_rows()
    write_grouped_workbook(headers, rows)
    print(
        f"Generated grouped API workbook: {OUTPUT_XLSX} "
        f"with {len(rows)} endpoints"
    )


if __name__ == "__main__":
    main()


