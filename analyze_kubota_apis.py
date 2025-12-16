import os
import re
from pathlib import Path
from typing import Dict, List, Set, Tuple
import openpyxl
from openpyxl import load_workbook

class KubotaAPIAnalyzer:
    def __init__(self, backend_path: str, excel_path: str):
        self.backend_path = Path(backend_path)
        self.excel_path = excel_path
        self.api_info = {}
        self.db_config = {}
        
    def find_application_properties(self) -> Dict[str, str]:
        """Find database configuration from application.properties"""
        props_files = list(self.backend_path.rglob("application*.properties"))
        props_files.extend(list(self.backend_path.rglob("application*.yml")))
        
        db_info = {}
        for props_file in props_files:
            try:
                with open(props_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                    # Extract database URL
                    db_url_match = re.search(r'spring\.datasource\.url\s*=\s*(.+)', content)
                    if db_url_match:
                        db_url = db_url_match.group(1).strip()
                        db_info['database_url'] = db_url
                        # Extract database name
                        db_name_match = re.search(r'jdbc:.*?://.*?/([^?;]+)', db_url)
                        if db_name_match:
                            db_info['database_name'] = db_name_match.group(1)
                    
                    # Extract driver
                    driver_match = re.search(r'spring\.datasource\.driver-class-name\s*=\s*(.+)', content)
                    if driver_match:
                        db_info['driver'] = driver_match.group(1).strip()
            except Exception as e:
                print(f"Error reading {props_file}: {e}")
        
        return db_info
    
    def parse_controller_file(self, file_path: Path) -> List[Dict]:
        """Parse a controller file to extract API endpoints"""
        endpoints = []
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            # Get base mapping
            base_mapping_match = re.search(r'@RequestMapping\s*\([^)]*value\s*=\s*["\']([^"\']+)["\']', content)
            base_path = base_mapping_match.group(1) if base_mapping_match else ""
            
            # Get controller class name
            class_match = re.search(r'public\s+class\s+(\w+)\s+', content)
            controller_name = class_match.group(1) if class_match else ""
            
            # Find all autowired repositories/services
            autowired_repos = re.findall(r'@Autowired\s+private\s+(\w+)\s+(\w+);', content)
            
            # Find all endpoint mappings with better pattern matching
            endpoint_patterns = [
                (r'@GetMapping\s*\([^)]*value\s*=\s*["\']([^"\']+)["\']', 'GET'),
                (r'@PostMapping\s*\([^)]*value\s*=\s*["\']([^"\']+)["\']', 'POST'),
                (r'@PutMapping\s*\([^)]*value\s*=\s*["\']([^"\']+)["\']', 'PUT'),
                (r'@DeleteMapping\s*\([^)]*value\s*=\s*["\']([^"\']+)["\']', 'DELETE'),
                (r'@RequestMapping\s*\([^)]*value\s*=\s*["\']([^"\']+)["\']', 'REQUEST'),
            ]
            
            for pattern, method in endpoint_patterns:
                matches = re.finditer(pattern, content)
                for match in matches:
                    endpoint_path = match.group(1)
                    # Clean up path
                    if not endpoint_path.startswith('/'):
                        endpoint_path = '/' + endpoint_path
                    full_path = base_path.rstrip('/') + endpoint_path if base_path else endpoint_path
                    
                    # Find the method after this annotation
                    method_start = match.end()
                    # Look for the next public method
                    method_search = content[method_start:method_start+1000]
                    method_match = re.search(r'public\s+\S+\s+(\w+)\s*\(', method_search)
                    method_name = method_match.group(1) if method_match else ""
                    
                    # Extract method body more accurately
                    method_body = self.extract_method_body(content, method_start)
                    
                    # Find repository/service calls in method body
                    used_repos = []
                    used_methods = []
                    
                    for repo_type, repo_name in autowired_repos:
                        if repo_name in method_body:
                            used_repos.append((repo_type, repo_name))
                            # Find methods called on this repo
                            repo_method_pattern = rf'{repo_name}\.(\w+)\s*\('
                            repo_methods = re.findall(repo_method_pattern, method_body)
                            for rm in repo_methods:
                                used_methods.append(f"{repo_name}.{rm}()")
                    
                    endpoints.append({
                        'endpoint': full_path,
                        'method': method,
                        'controller_file': str(file_path.relative_to(self.backend_path)),
                        'controller_class': controller_name,
                        'method_name': method_name,
                        'repositories': used_repos,
                        'repo_methods': used_methods
                    })
        
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
        
        return endpoints
    
    def extract_method_body(self, content: str, start_pos: int) -> str:
        """Extract method body from a given position"""
        # Find the next public method
        method_search = content[start_pos:start_pos+1000]
        method_match = re.search(r'public\s+[^{]*\{', method_search)
        if not method_match:
            return ""
        
        # Find the opening brace
        brace_pos = start_pos + method_match.end() - 1
        brace_count = 1
        pos = brace_pos + 1
        
        while brace_count > 0 and pos < len(content):
            if content[pos] == '{':
                brace_count += 1
            elif content[pos] == '}':
                brace_count -= 1
            pos += 1
        
        return content[brace_pos:pos]
    
    def find_repository_files(self, repo_name: str) -> List[Path]:
        """Find repository interface files"""
        repo_files = []
        # Clean repo name
        clean_name = repo_name.replace('Repo', '').replace('Repository', '')
        
        # Try multiple patterns
        patterns = [
            f"*{repo_name}*.java",
            f"*{clean_name}*Repository.java",
            f"*{clean_name}*Repo.java",
        ]
        
        for pattern in patterns:
            repo_files.extend(list(self.backend_path.rglob(pattern)))
        
        # Remove duplicates
        return list(set(repo_files))
    
    def find_entity_files(self, repo_name: str) -> List[Path]:
        """Find entity/domain files related to repository"""
        entity_files = []
        # Extract entity name from repository name
        entity_name = repo_name.replace('Repo', '').replace('Repository', '')
        
        # Find domain/entity files
        pattern = f"*{entity_name}*.java"
        entity_files.extend(list(self.backend_path.rglob(pattern)))
        
        return entity_files
    
    def analyze_repository(self, repo_name: str) -> Dict:
        """Analyze a repository to find entity and database info"""
        repo_info = {
            'repository_file': '',
            'entity_file': '',
            'entity_name': '',
            'db_object_type': 'Repository'
        }
        
        repo_files = self.find_repository_files(repo_name)
        if repo_files:
            repo_info['repository_file'] = str(repo_files[0].relative_to(self.backend_path))
            
            # Read repository file to find entity
            try:
                with open(repo_files[0], 'r', encoding='utf-8') as f:
                    repo_content = f.read()
                    # Find extends JpaRepository<Entity, ID> or similar
                    entity_patterns = [
                        r'extends\s+\w+Repository\s*<\s*(\w+)',
                        r'extends\s+JpaRepository\s*<\s*(\w+)',
                        r'extends\s+CrudRepository\s*<\s*(\w+)',
                        r'extends\s+PagingAndSortingRepository\s*<\s*(\w+)',
                    ]
                    
                    entity_name = None
                    for pattern in entity_patterns:
                        entity_match = re.search(pattern, repo_content)
                        if entity_match:
                            entity_name = entity_match.group(1)
                            break
                    
                    if entity_name:
                        repo_info['entity_name'] = entity_name
                        
                        # Find entity file
                        entity_files = self.find_entity_files(entity_name)
                        if entity_files:
                            repo_info['entity_file'] = str(entity_files[0].relative_to(self.backend_path))
                            
                            # Check if it's an entity and get table name
                            with open(entity_files[0], 'r', encoding='utf-8') as ef:
                                entity_content = ef.read()
                                if '@Entity' in entity_content:
                                    repo_info['db_object_type'] = 'JPA Entity'
                                    # Try to find table name
                                    table_match = re.search(r'@Table\s*\([^)]*name\s*=\s*["\']([^"\']+)["\']', entity_content)
                                    if table_match:
                                        repo_info['table_name'] = table_match.group(1)
            except Exception as e:
                print(f"Error analyzing repository {repo_name}: {e}")
        
        return repo_info
    
    def analyze_all_controllers(self):
        """Analyze all controller files"""
        controller_files = list(self.backend_path.rglob("*Controller.java"))
        print(f"Found {len(controller_files)} controller files")
        
        all_endpoints = []
        for idx, controller_file in enumerate(controller_files, 1):
            if idx % 10 == 0:
                print(f"Processing controller {idx}/{len(controller_files)}...")
            endpoints = self.parse_controller_file(controller_file)
            all_endpoints.extend(endpoints)
        
        print(f"Found {len(all_endpoints)} total endpoints")
        return all_endpoints
    
    def get_service_files(self, controller_file: str) -> List[str]:
        """Find service files used by controller"""
        service_files = []
        controller_path = self.backend_path / controller_file
        
        try:
            with open(controller_path, 'r', encoding='utf-8') as f:
                content = f.read()
                # Find @Autowired services
                service_matches = re.findall(r'@Autowired\s+private\s+(\w+)\s+(\w+);', content)
                for service_type, service_name in service_matches:
                    if 'Service' in service_type:
                        # Find service file - try multiple patterns
                        patterns = [
                            f"*{service_type}.java",
                            f"*{service_name}.java",
                        ]
                        for pattern in patterns:
                            service_files_found = list(self.backend_path.rglob(pattern))
                            if service_files_found:
                                for sf in service_files_found:
                                    service_file_str = str(sf.relative_to(self.backend_path))
                                    if service_file_str not in service_files:
                                        service_files.append(service_file_str)
                                break
        except Exception as e:
            print(f"Error finding services for {controller_file}: {e}")
        
        return service_files
    
    def update_excel(self):
        """Read Excel, analyze APIs, and update with new columns"""
        # Load existing Excel
        wb = load_workbook(self.excel_path)
        ws = wb.active
        
        # Get database config
        self.db_config = self.find_application_properties()
        db_name = self.db_config.get('database_name', 'Unknown')
        
        # Analyze all controllers
        print("Analyzing controllers...")
        all_endpoints = self.analyze_all_controllers()
        
        # Create a mapping of endpoints
        endpoint_map = {}
        for ep in all_endpoints:
            key = f"{ep['method']} {ep['endpoint']}"
            if key not in endpoint_map:
                endpoint_map[key] = []
            endpoint_map[key].append(ep)
        
        # Read existing data
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Add new columns if they don't exist
        new_columns = ['Files Used', 'Methods Used', 'Database Used', 'DB Objects Used', 'DB Object Type']
        for col in new_columns:
            if col not in headers:
                headers.append(col)
                ws.cell(row=1, column=len(headers), value=col)
        
        # Update each row
        print(f"Updating {ws.max_row - 1} rows in Excel...")
        for row_idx in range(2, ws.max_row + 1):
            if (row_idx - 1) % 50 == 0:
                print(f"Processing row {row_idx - 1}/{ws.max_row - 1}...")
            # Try to find endpoint in row
            endpoint_cell = None
            method_cell = None
            
            # Look for endpoint URL in various columns
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    # Check for API endpoint
                    if '/api/' in cell_str or cell_str.startswith('/api/') or cell_str.startswith('api/'):
                        endpoint_cell = cell_str
                    # Check for HTTP method
                    if cell_str.upper() in ['GET', 'POST', 'PUT', 'DELETE', 'PATCH']:
                        method_cell = cell_str.upper()
            
            if endpoint_cell:
                # Normalize endpoint
                if not endpoint_cell.startswith('/'):
                    endpoint_cell = '/' + endpoint_cell
                if not endpoint_cell.startswith('/api'):
                    endpoint_cell = '/api' + endpoint_cell
                
                # Try to find matching endpoint (with or without method)
                matching_endpoints = []
                if method_cell:
                    search_key = f"{method_cell} {endpoint_cell}"
                    matching_endpoints = endpoint_map.get(search_key, [])
                
                # If no match with method, try without method
                if not matching_endpoints:
                    for key, eps in endpoint_map.items():
                        if endpoint_cell in key:
                            matching_endpoints.extend(eps)
                
                if matching_endpoints:
                    ep = matching_endpoints[0]
                    
                    # Collect all files
                    files_used = [ep['controller_file']]
                    
                    # Get service files
                    service_files = self.get_service_files(ep['controller_file'])
                    files_used.extend(service_files)
                    
                    # Get repository and entity files
                    repo_files_list = []
                    entity_files_list = []
                    db_objects = []
                    db_object_types = []
                    
                    for repo_type, repo_name in ep['repositories']:
                        repo_info = self.analyze_repository(repo_name)
                        if repo_info['repository_file']:
                            repo_files_list.append(repo_info['repository_file'])
                            db_objects.append(repo_name)
                            db_object_types.append(repo_info['db_object_type'])
                        if repo_info['entity_file']:
                            entity_files_list.append(repo_info['entity_file'])
                    
                    files_used.extend(repo_files_list)
                    files_used.extend(entity_files_list)
                    
                    # Get methods used
                    methods_used = [ep['method_name']] if ep['method_name'] else []
                    methods_used.extend(ep.get('repo_methods', []))
                    
                    # Update cells
                    files_col = headers.index('Files Used') + 1 if 'Files Used' in headers else len(headers) + 1
                    methods_col = headers.index('Methods Used') + 1 if 'Methods Used' in headers else len(headers) + 2
                    db_col = headers.index('Database Used') + 1 if 'Database Used' in headers else len(headers) + 3
                    db_obj_col = headers.index('DB Objects Used') + 1 if 'DB Objects Used' in headers else len(headers) + 4
                    db_type_col = headers.index('DB Object Type') + 1 if 'DB Object Type' in headers else len(headers) + 5
                    
                    ws.cell(row=row_idx, column=files_col, value='; '.join(set(files_used)))
                    ws.cell(row=row_idx, column=methods_col, value='; '.join(set(methods_used)))
                    ws.cell(row=row_idx, column=db_col, value=db_name)
                    ws.cell(row=row_idx, column=db_obj_col, value='; '.join(set(db_objects)))
                    ws.cell(row=row_idx, column=db_type_col, value='; '.join(set(db_object_types)))
        
        # Save updated Excel
        wb.save(self.excel_path)
        print(f"Updated Excel file: {self.excel_path}")

if __name__ == "__main__":
    backend_path = "KUBOTA/KUBOTA-BACKENED"
    excel_path = "KUBOTA_API_Endpoints.xlsx"
    
    analyzer = KubotaAPIAnalyzer(backend_path, excel_path)
    analyzer.update_excel()



