import csv
import os
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STRUTS_CONFIG = os.path.join(
    BASE_DIR, "src", "main", "webapp", "WEB-INF", "struts-config.xml"
)
WEB_XML = os.path.join(
    BASE_DIR, "src", "main", "webapp", "WEB-INF", "web.xml"
)
OUTPUT_CSV = os.path.join(BASE_DIR, "ITLDIS_API_Endpoints_Summary.csv")
OUTPUT_XLSX = os.path.join(BASE_DIR, "ITLDIS_API_Endpoints_Detailed.xlsx")


def infer_module_from_path(path: str, action_type: str | None) -> str:
    """
    Very lightweight heuristic to group APIs by module.
    """
    if path.startswith("/camunda/"):
        return "Camunda BPM"
    if path.startswith("/inventoryEXP"):
        return "Inventory Export"
    if path.startswith("/inventory"):
        return "Inventory"
    if path.startswith("/warranty"):
        return "Warranty"
    if path.startswith("/Service") or path.startswith("/service"):
        return "Service"
    if path.startswith("/pdi"):
        return "PDI"
    if path.startswith("/install"):
        return "Installation"
    if path.startswith("/UserManagement"):
        return "User Management"
    if path.startswith("/manageCustomer"):
        return "Customer Management"
    if path.startswith("/reportAction"):
        return "Reports"
    if path.startswith("/EAMG_") or (action_type and action_type.startswith("EAMG.")):
        return "EAMG Catalogue"
    if path.startswith("/webservice"):
        return "External SAP Webservice"
    if path.startswith("/login") or path.startswith("/ChangeLogin") or path.startswith("/Welcome"):
        return "Authentication"
    if path.startswith("/DigitalSignatureAction"):
        return "Digital Signature"
    return "General"


def load_struts_actions():
    tree = ET.parse(STRUTS_CONFIG)
    root = tree.getroot()
    ns = {}  # no namespace

    actions = []
    for action in root.findall(".//action-mappings/action", ns):
        path = action.get("path")
        if not path:
            continue

        # Struts actions are served via *.do mapping
        endpoint = f"{path}.do"
        action_type = action.get("type")

        module = infer_module_from_path(path, action_type)

        # Request/response description – high level
        request_desc = "Form parameters (Struts form / query parameters)"
        response_desc = "HTML (JSP view) via Struts forward"

        # Implementation / DB notes – generic but helpful
        impl_files = []
        if action_type:
            impl_files.append(action_type.replace(".", "/") + ".java")
        name = action.get("name")
        if name:
            impl_files.append(f"FormBean: {name}")

        impl_files_str = "; ".join(impl_files) if impl_files else ""

        actions.append(
            {
                "Module": module,
                "API Endpoint": endpoint,
                "HTTP Method": "POST",  # Struts actions typically handle POST (and often GET) via doPost/doGet
                "Request": request_desc,
                "Response": response_desc,
                "User-Defined Files": impl_files_str,
                "User-Defined Methods": "Dispatch methods / execute() inside Action class (varies by 'option' or 'method' parameter)",
                "Database": "ITLDIS (SQL Server)",
                "DB Object": "Access via DAO classes and Hibernate entities (see 'dao' and 'HibernateMapping' packages)",
                "DB Object Type": "Tables / Views / Stored Procedures (via Hibernate/JDBC, endpoint-specific)",
            }
        )

    return actions


def load_servlet_endpoints():
    tree = ET.parse(WEB_XML)
    root = tree.getroot()

    # web.xml is namespace-qualified; grab the default namespace
    ns_uri = root.tag.split("}")[0].strip("{") if "}" in root.tag else ""
    ns = {"ns": ns_uri} if ns_uri else {}

    servlets = {}
    for servlet in root.findall("ns:servlet", ns) if ns else root.findall("servlet"):
        name_el = servlet.find("ns:servlet-name", ns) if ns else servlet.find("servlet-name")
        class_el = servlet.find("ns:servlet-class", ns) if ns else servlet.find("servlet-class")
        if name_el is None:
            continue
        servlets[name_el.text] = class_el.text if class_el is not None else None

    endpoints = []
    for mapping in root.findall("ns:servlet-mapping", ns) if ns else root.findall("servlet-mapping"):
        name_el = mapping.find("ns:servlet-name", ns) if ns else mapping.find("servlet-name")
        pattern_el = mapping.find("ns:url-pattern", ns) if ns else mapping.find("url-pattern")
        if name_el is None or pattern_el is None:
            continue

        url_pattern = pattern_el.text or ""

        # Skip wildcard infrastructure mappings, focus on concrete endpoints
        if url_pattern.endswith("*.do") or url_pattern.endswith("*.jsp"):
            continue

        servlet_name = name_el.text
        servlet_class = servlets.get(servlet_name)

        module = infer_module_from_path(url_pattern, servlet_class)

        request_desc = "HTTP request (typically query parameters / form data)"
        response_desc = "HTML (JSP) or file download depending on servlet"

        impl_files = []
        if servlet_class:
            impl_files.append(servlet_class.replace(".", "/") + ".java")

        impl_files_str = "; ".join(impl_files) if impl_files else ""

        endpoints.append(
            {
                "Module": module,
                "API Endpoint": url_pattern,
                "HTTP Method": "GET/POST",
                "Request": request_desc,
                "Response": response_desc,
                "User-Defined Files": impl_files_str,
                "User-Defined Methods": "doGet()/doPost() and helper methods inside servlet",
                "Database": "ITLDIS (SQL Server)",
                "DB Object": "Access via DAO/utility classes and Hibernate entities (servlet-specific)",
                "DB Object Type": "Tables / Views / Stored Procedures (via Hibernate/JDBC, endpoint-specific)",
            }
        )

    return endpoints


def main():
    fieldnames = [
        "Module",
        "API Endpoint",
        "HTTP Method",
        "Request",
        "Response",
        "User-Defined Files",
        "User-Defined Methods",
        "Database",
        "DB Object",
        "DB Object Type",
    ]

    actions = load_struts_actions()
    servlets = load_servlet_endpoints()

    all_rows = actions + servlets

    # Sort by module then endpoint for readability
    all_rows.sort(key=lambda r: (r["Module"], r["API Endpoint"]))

    # Try to write CSV; if it's locked (e.g. open in Excel), skip rewriting it.
    try:
        with open(OUTPUT_CSV, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in all_rows:
                writer.writerow(row)
        csv_written = True
    except PermissionError:
        csv_written = False

    # Write detailed Excel (XLSX)
    wb = Workbook()
    ws = wb.active
    ws.title = "API Endpoints"

    # Header row with bold font
    header_font = Font(bold=True)
    ws.append(fieldnames)
    for col_idx in range(1, len(fieldnames) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font

    # Data rows
    for row in all_rows:
        ws.append([row.get(col, "") for col in fieldnames])

    # Auto-size columns
    for col_idx, col_name in enumerate(fieldnames, start=1):
        max_length = len(col_name)
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    # Freeze header and add filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_XLSX)

    print(
        f"Generated {len(all_rows)} API endpoint rows into {OUTPUT_CSV} "
        f"and {OUTPUT_XLSX}"
    )


if __name__ == "__main__":
    main()


