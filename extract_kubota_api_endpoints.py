import os
import re
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Standard Java/Spring library classes to exclude
STANDARD_LIBRARIES = {
    'System', 'String', 'Integer', 'Long', 'Double', 'Boolean', 'Date', 'List', 'Map', 'Set',
    'ArrayList', 'HashMap', 'HashSet', 'Collections', 'Arrays', 'Objects',
    'ResponseEntity', 'HttpStatus', 'RequestMethod', 'PathVariable', 'RequestParam', 'RequestBody',
    'Autowired', 'RestController', 'RequestMapping', 'GetMapping', 'PostMapping', 'PutMapping',
    'DeleteMapping', 'PatchMapping', 'CrossOrigin', 'Valid', 'Exception', 'RuntimeException',
    'Logger', 'LoggerFactory', 'Optional', 'Stream', 'Collectors'
}

def is_user_defined(class_name):
    """Check if a class is user-defined (not a standard library)"""
    if not class_name:
        return False
    # Remove generics
    class_name = re.sub(r'<.*>', '', class_name).strip()
    # Check if it's a standard library
    if class_name in STANDARD_LIBRARIES:
        return False
    # Check if it starts with java. or javax. or org.springframework (except our packages)
    if class_name.startswith('java.') or class_name.startswith('javax.') or class_name.startswith('org.springframework'):
        return False
    return True

def find_entity_file(entity_name, base_path):
    """Find entity file given entity class name"""
    patterns = [
        f"**/{entity_name}.java",
        f"**/domain/{entity_name}.java",
    ]
    
    for pattern in patterns:
        for file_path in base_path.rglob(pattern):
            if 'domain' in str(file_path) or 'entity' in str(file_path).lower():
                return str(file_path)
    
    return None

def extract_table_from_entity(entity_file_path):
    """Extract table name from entity file"""
    try:
        if not Path(entity_file_path).exists():
            return None
        
        with open(entity_file_path, 'r', encoding='utf-8', errors='ignore') as f:
            entity_content = f.read()
        
        # Extract table name from @Table annotation
        table_pattern = r'@Table\s*\(\s*name\s*=\s*["\']([^"\']+)["\']'
        table_match = re.search(table_pattern, entity_content)
        if table_match:
            return table_match.group(1)
        
        # If no @Table annotation, use class name converted to snake_case
        class_pattern = r'public\s+class\s+(\w+)'
        class_match = re.search(class_pattern, entity_content)
        if class_match:
            class_name = class_match.group(1)
            # Convert CamelCase to snake_case
            table_name = re.sub(r'(?<!^)(?=[A-Z])', '_', class_name).upper()
            return table_name
        
    except Exception:
        pass
    
    return None

def extract_db_objects_from_repository(repo_file_path, base_path):
    """Extract database objects (tables, stored procedures, etc.) from repository file"""
    db_objects = []
    db_object_types = []
    
    try:
        if not Path(repo_file_path).exists():
            return db_objects, db_object_types
        
        with open(repo_file_path, 'r', encoding='utf-8', errors='ignore') as f:
            repo_content = f.read()
        
        # Extract stored procedures from @Query annotations
        # Pattern: @Query(value = "{call sp_procedure_name(...)}", nativeQuery = true)
        sp_pattern = r'@Query\s*\([^)]*value\s*=\s*["\']\{call\s+(\w+)\s*\([^)]*\)\}["\']'
        stored_procs = re.findall(sp_pattern, repo_content, re.IGNORECASE)
        for sp in stored_procs:
            if sp not in db_objects:
                db_objects.append(sp)
                db_object_types.append("Stored Procedure")
        
        # Also extract from direct {call ...} patterns
        sp_pattern2 = r'\{call\s+(\w+)\s*\([^)]*\)\}'
        stored_procs2 = re.findall(sp_pattern2, repo_content, re.IGNORECASE)
        for sp in stored_procs2:
            if (sp.startswith('sp_') or sp.startswith('SP_')) and sp not in db_objects:
                db_objects.append(sp)
                db_object_types.append("Stored Procedure")
        
        # Extract entity class name from JpaRepository<EntityName, ...>
        entity_pattern = r'extends\s+JpaRepository<(\w+)'
        entity_matches = re.findall(entity_pattern, repo_content)
        for entity_name in entity_matches:
            # Find entity file and extract table name
            entity_file = find_entity_file(entity_name, base_path)
            if entity_file:
                table_name = extract_table_from_entity(entity_file)
                if table_name and table_name not in db_objects:
                    db_objects.append(table_name)
                    db_object_types.append("Table")
        
        # Extract table names from native queries (SELECT FROM table_name)
        # Look for @Query with nativeQuery = true
        native_query_pattern = r'@Query\s*\([^)]*nativeQuery\s*=\s*true[^)]*value\s*=\s*["\']([^"\']+)["\']'
        native_queries = re.findall(native_query_pattern, repo_content, re.IGNORECASE | re.DOTALL)
        for query in native_queries:
            # Extract table names from FROM, JOIN, UPDATE, INSERT INTO
            table_pattern = r'(?:FROM|JOIN|UPDATE|INTO)\s+(\w+)'
            tables = re.findall(table_pattern, query, re.IGNORECASE)
            for table in tables:
                if table.upper() not in ['SELECT', 'WHERE', 'AND', 'OR', 'ORDER', 'GROUP', 'HAVING', 'SET', 'VALUES']:
                    if table not in db_objects:
                        db_objects.append(table)
                        db_object_types.append("Table (Native Query)")
        
        # Extract views (tables ending with _view or View)
        view_pattern = r'(?:FROM|JOIN)\s+(\w+_view|\w+View)\s+'
        views = re.findall(view_pattern, repo_content, re.IGNORECASE)
        for view in views:
            if view not in db_objects:
                db_objects.append(view)
                db_object_types.append("View")
    
    except Exception as e:
        pass
    
    return db_objects, db_object_types

def find_repository_file(repo_name, base_path):
    """Find the repository file given a repository name"""
    # Try different patterns
    patterns = [
        f"**/{repo_name}.java",
        f"**/*{repo_name}.java",
    ]
    
    for pattern in patterns:
        for file_path in base_path.rglob(pattern):
            if 'repository' in str(file_path).lower():
                return str(file_path)
    
    return None

def extract_user_defined_methods(content, method_start_line, base_package, base_path):
    """Extract user-defined methods called within a method"""
    lines = content.split('\n')
    method_lines = []
    brace_count = 0
    in_method = False
    
    for i in range(method_start_line, len(lines)):
        line = lines[i]
        method_lines.append(line)
        
        brace_count += line.count('{') - line.count('}')
        if '{' in line:
            in_method = True
        if in_method and brace_count == 0 and '}' in line:
            break
    
    method_content = '\n'.join(method_lines)
    
    # Extract request type
    request_match = re.search(r'@RequestBody\s+(?:final\s+)?(\w+<?[^>]*>?)', method_content)
    request_type = request_match.group(1) if request_match else ""
    
    # Extract response type
    response_match = re.search(r'public\s+(?:ResponseEntity<)?(\w+<?[^>]*>?)(?:>)?\s+\w+', method_content)
    response_type = response_match.group(1) if response_match else ""
    
    # Extract user-defined repository/service calls
    user_repos = []
    user_services = []
    user_methods = []
    all_db_objects = []
    all_db_object_types = []
    
    # Find repository calls (user-defined repositories)
    repo_pattern = r'(\w+Repo(?:sitory)?)\.(\w+)\s*\('
    repo_matches = re.findall(repo_pattern, method_content)
    for repo_name, method_name in repo_matches:
        if is_user_defined(repo_name):
            user_repos.append(f"{repo_name}.{method_name}")
            user_methods.append(f"{repo_name}.{method_name}")
            
            # Find the repository file and extract DB objects
            repo_file = find_repository_file(repo_name, base_path)
            if repo_file:
                db_objs, db_types = extract_db_objects_from_repository(repo_file, base_path)
                all_db_objects.extend(db_objs)
                all_db_object_types.extend(db_types)
    
    # Find service calls (user-defined services)
    service_pattern = r'(\w+Service)\.(\w+)\s*\('
    service_matches = re.findall(service_pattern, method_content)
    for service_name, method_name in service_matches:
        if is_user_defined(service_name):
            user_services.append(f"{service_name}.{method_name}")
            user_methods.append(f"{service_name}.{method_name}")
    
    # Find other user-defined method calls
    method_call_pattern = r'(\w+)\.(\w+)\s*\('
    all_calls = re.findall(method_call_pattern, method_content)
    
    for obj_name, method_name in all_calls:
        if obj_name in STANDARD_LIBRARIES or obj_name.startswith('apiResponse') or obj_name.startswith('logger'):
            continue
        if obj_name[0].isupper() and is_user_defined(obj_name):
            if f"{obj_name}.{method_name}" not in user_methods:
                user_methods.append(f"{obj_name}.{method_name}")
    
    # Extract stored procedures directly from method content
    sp_pattern = r'\{call\s+(\w+)\s*\([^)]*\)\}'
    stored_procs = re.findall(sp_pattern, method_content, re.IGNORECASE)
    for sp in stored_procs:
        if (sp.startswith('sp_') or sp.startswith('SP_')) and sp not in all_db_objects:
            all_db_objects.append(sp)
            all_db_object_types.append("Stored Procedure")
    
    # Extract database type
    db_type = ""
    if any('Repository' in r for r in user_repos):
        db_type = "JPA Repository"
    elif 'EntityManager' in method_content:
        db_type = "JPA EntityManager"
    elif 'JdbcTemplate' in method_content:
        db_type = "JDBC Template"
    
    return {
        'request_type': request_type,
        'response_type': response_type,
        'user_repos': user_repos,
        'user_services': user_services,
        'user_methods': user_methods,
        'db_objects': list(set(all_db_objects)),  # Remove duplicates
        'db_object_types': list(set(all_db_object_types)),  # Remove duplicates
        'db_type': db_type
    }

def find_user_defined_files(controller_file, dependencies, content):
    """Find user-defined files used (services, repositories, DTOs, domains)"""
    user_files = []
    
    # Add controller file
    try:
        controller_file_str = str(controller_file.relative_to(Path.cwd()))
    except ValueError:
        controller_file_str = str(controller_file)
    user_files.append(controller_file_str)
    
    # Find imports to identify user-defined classes
    import_pattern = r'import\s+com\.i4o\.dms\.kubota\.([^;]+);'
    imports = re.findall(import_pattern, content)
    
    for imp in imports:
        # Extract class name
        class_name = imp.split('.')[-1]
        if is_user_defined(class_name):
            # Try to find the file
            parts = imp.split('.')
            if len(parts) >= 2:
                # Construct potential file path
                package_path = '/'.join(parts[:-1])
                file_name = parts[-1] + '.java'
                potential_path = f"KUBOTA/KUBOTA-BACKENED/src/main/java/com/i4o/dms/kubota/{package_path}/{file_name}"
                if Path(potential_path).exists():
                    user_files.append(potential_path)
    
    # Add dependencies that are user-defined
    for dep_name, dep_type in dependencies.items():
        if is_user_defined(dep_type):
            user_files.append(f"{dep_type} ({dep_name})")
    
    return list(set(user_files))  # Remove duplicates

def extract_api_endpoints():
    """Extract all API endpoints from KUBOTA backend"""
    
    base_path = Path("KUBOTA/KUBOTA-BACKENED/src/main/java")
    controllers = []
    
    # Find all controller files
    for controller_file in base_path.rglob("*Controller.java"):
        controllers.append(controller_file)
    
    endpoints = []
    
    for controller_file in controllers:
        try:
            with open(controller_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                lines = content.split('\n')
                
            # Extract base package
            package_match = re.search(r'package\s+(com\.i4o\.dms\.kubota[^;]+);', content)
            base_package = package_match.group(1) if package_match else ""
            
            # Extract class-level RequestMapping
            class_mapping_match = re.search(r'@RequestMapping\s*\([^)]*value\s*=\s*["\']([^"\']+)["\']', content)
            base_path_api = class_mapping_match.group(1) if class_mapping_match else ""
            
            # Extract RestController annotation
            is_rest = '@RestController' in content
            
            if not is_rest:
                continue
            
            # Extract @Autowired dependencies (user-defined)
            autowired_deps = re.findall(r'@Autowired\s+private\s+(\w+)\s+(\w+)', content)
            dependencies = {}
            for dep_type, dep_name in autowired_deps:
                if is_user_defined(dep_type):
                    dependencies[dep_name] = dep_type
            
            # Find all method mappings
            for i, line in enumerate(lines):
                http_method = None
                endpoint_path = None
                method_name = None
                
                # Check for various mapping annotations
                if '@GetMapping' in line:
                    match = re.search(r'@GetMapping\s*\([^)]*["\']([^"\']+)["\']', line)
                    if match:
                        endpoint_path = match.group(1)
                        http_method = 'GET'
                elif '@PostMapping' in line:
                    match = re.search(r'@PostMapping\s*\([^)]*["\']([^"\']+)["\']', line)
                    if match:
                        endpoint_path = match.group(1)
                        http_method = 'POST'
                elif '@PutMapping' in line:
                    match = re.search(r'@PutMapping\s*\([^)]*["\']([^"\']+)["\']', line)
                    if match:
                        endpoint_path = match.group(1)
                        http_method = 'PUT'
                elif '@DeleteMapping' in line:
                    match = re.search(r'@DeleteMapping\s*\([^)]*["\']([^"\']+)["\']', line)
                    if match:
                        endpoint_path = match.group(1)
                        http_method = 'DELETE'
                elif '@PatchMapping' in line:
                    match = re.search(r'@PatchMapping\s*\([^)]*["\']([^"\']+)["\']', line)
                    if match:
                        endpoint_path = match.group(1)
                        http_method = 'PATCH'
                elif '@RequestMapping' in line and ('method' in line.lower() or 'value' in line.lower()):
                    path_match = re.search(r'value\s*=\s*["\']([^"\']+)["\']', line)
                    method_match = re.search(r'method\s*=\s*RequestMethod\.(\w+)', line)
                    if path_match:
                        endpoint_path = path_match.group(1)
                        http_method = method_match.group(1) if method_match else 'REQUEST'
                
                if endpoint_path and http_method:
                    # Find the method definition (look ahead up to 15 lines)
                    for j in range(i+1, min(i+15, len(lines))):
                        method_def_match = re.search(r'public\s+(?:ResponseEntity<)?\w+<?[^>]*>?\s+(\w+)\s*\(', lines[j])
                        if method_def_match:
                            method_name = method_def_match.group(1)
                            method_start = j
                            
                            # Extract method details
                            method_details = extract_user_defined_methods(content, method_start, base_package, base_path)
                            
                            # Build full endpoint path
                            full_path = base_path_api + endpoint_path if base_path_api else endpoint_path
                            if not full_path.startswith('/'):
                                full_path = '/' + full_path
                            
                            # Get user-defined files
                            user_files = find_user_defined_files(controller_file, dependencies, content)
                            
                            try:
                                controller_file_str = str(controller_file.relative_to(Path.cwd()))
                            except ValueError:
                                controller_file_str = str(controller_file)
                            
                            endpoint = {
                                'api_endpoint': full_path,
                                'http_method': http_method,
                                'request': method_details['request_type'],
                                'response': method_details['response_type'],
                                'when_called': method_name,
                                'controller_file': controller_file_str,
                                'method_name': method_name,
                                'user_files_used': '; '.join(user_files[:20]),  # Limit to first 20
                                'user_methods_used': '; '.join(method_details['user_methods'][:20]),  # Limit to first 20
                                'database_used': method_details['db_type'],
                                'user_db_objects': '; '.join(method_details['db_objects']),
                                'db_object_type': '; '.join(method_details['db_object_types']) if method_details['db_object_types'] else method_details['db_type']
                            }
                            
                            endpoints.append(endpoint)
                            break
            
        except Exception as e:
            continue
    
    return endpoints

def create_excel(endpoints):
    """Create Excel file with API endpoint details"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "KUBOTA API Endpoints"
    
    # Headers
    headers = [
        'API Endpoint',
        'HTTP Method',
        'Request Type',
        'Response Type',
        'When Called (Method Name)',
        'Controller File',
        'Method Name',
        'User-Defined Files Used',
        'User-Defined Methods Used',
        'Database Used',
        'User-Defined DB Objects (Tables, Stored Procedures, etc.)',
        'DB Object Type'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Add data
    for row_num, endpoint in enumerate(endpoints, 2):
        ws.cell(row=row_num, column=1, value=endpoint.get('api_endpoint', ''))
        ws.cell(row=row_num, column=2, value=endpoint.get('http_method', ''))
        ws.cell(row=row_num, column=3, value=endpoint.get('request', ''))
        ws.cell(row=row_num, column=4, value=endpoint.get('response', ''))
        ws.cell(row=row_num, column=5, value=endpoint.get('when_called', ''))
        ws.cell(row=row_num, column=6, value=endpoint.get('controller_file', ''))
        ws.cell(row=row_num, column=7, value=endpoint.get('method_name', ''))
        ws.cell(row=row_num, column=8, value=endpoint.get('user_files_used', ''))
        ws.cell(row=row_num, column=9, value=endpoint.get('user_methods_used', ''))
        ws.cell(row=row_num, column=10, value=endpoint.get('database_used', ''))
        ws.cell(row=row_num, column=11, value=endpoint.get('user_db_objects', ''))
        ws.cell(row=row_num, column=12, value=endpoint.get('db_object_type', ''))
        
        # Apply borders and wrap text
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    column_widths = {
        1: 45,  # API Endpoint
        2: 12,  # HTTP Method
        3: 25,  # Request Type
        4: 25,  # Response Type
        5: 25,  # When Called
        6: 50,  # Controller File
        7: 20,  # Method Name
        8: 60,  # User-Defined Files Used
        9: 50,  # User-Defined Methods Used
        10: 20, # Database Used
        11: 50, # User-Defined DB Objects
        12: 30  # DB Object Type
    }
    
    for col_num, width in column_widths.items():
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save file to KUBOTA folder
    output_file = "KUBOTA/KUBOTA_API_Endpoints_Detailed.xlsx"
    
    # Save directly to the output file
    try:
        wb.save(output_file)
        print(f"\nExcel file created: {output_file}")
        print(f"Total endpoints found: {len(endpoints)}")
    except Exception as e:
        # If file is open, save with timestamp
        import time
        timestamp = int(time.time())
        temp_file = f"KUBOTA/KUBOTA_API_Endpoints_Detailed_{timestamp}.xlsx"
        wb.save(temp_file)
        print(f"\nWarning: Could not save to {output_file} (file may be open in Excel)")
        print(f"File saved as: {temp_file}")
        print(f"Please close Excel and rename it to: {output_file}")
        output_file = temp_file
    
    return output_file

if __name__ == "__main__":
    print("Extracting KUBOTA API endpoints with database objects (tables, stored procedures, etc.)...")
    print("This may take a few minutes...")
    endpoints = extract_api_endpoints()
    
    if endpoints:
        print(f"\nFound {len(endpoints)} endpoints")
        create_excel(endpoints)
        print("\nDone!")
    else:
        print("No endpoints found!")
