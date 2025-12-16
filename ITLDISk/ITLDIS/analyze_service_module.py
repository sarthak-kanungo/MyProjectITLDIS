#!/usr/bin/env python3
"""
Script to analyze Service Module tiles, APIs, classes, methods, and database calls
Generates an Excel file with comprehensive analysis
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define the 13 Service Module Tiles based on dashboard.jsp
SERVICE_TILES = [
    {
        "tile_id": 1,
        "tile_name": "Pending for PDI",
        "functionality_code": "610",
        "action_path": "/pdiServiceProcess.do?option=init_viewallPendingChassisdetails&jobType=PDI",
        "action_method": "init_viewallPendingChassisdetails",
        "action_class": "action.pdiAction"
    },
    {
        "tile_id": 2,
        "tile_name": "View PDI Detail",
        "functionality_code": "605",
        "action_path": "/pdiServiceProcess.do?option=init_viewPDIDetails&jobType=PDI",
        "action_method": "init_viewPDIDetails",
        "action_class": "action.pdiAction"
    },
    {
        "tile_id": 3,
        "tile_name": "Pending Installation",
        "functionality_code": "611",
        "action_path": "/installProcess.do?option=initViewAllInstallList",
        "action_method": "initViewAllInstallList",
        "action_class": "action.installAction"
    },
    {
        "tile_id": 4,
        "tile_name": "View Installation",
        "functionality_code": "607",
        "action_path": "/installProcess.do?option=init_viewDetails",
        "action_method": "init_viewDetails",
        "action_class": "action.installAction"
    },
    {
        "tile_id": 5,
        "tile_name": "Create Job Card",
        "functionality_code": "608",
        "action_path": "/serviceCreateJC.do?option=initVehicleInformation&jobType=PDI&pathNm=fillJC",
        "action_method": "initVehicleInformation",
        "action_class": "action.createJobCardAction"
    },
    {
        "tile_id": 6,
        "tile_name": "View All Job Card",
        "functionality_code": "609",
        "action_path": "/serviceProcess.do?option=init_viewallJobcarddetails&jobType=ViewJC",
        "action_method": "init_viewallJobcarddetails",
        "action_class": "action.serviceAction"
    },
    {
        "tile_id": 7,
        "tile_name": "Close Job Card",
        "functionality_code": "616",
        "action_path": "/serviceProcess.do?option=initCloseJobCards&pathNm=fillJC",
        "action_method": "initCloseJobCards",
        "action_class": "action.serviceAction"
    },
    {
        "tile_id": 8,
        "tile_name": "Pending 4 Chassis Validate",
        "functionality_code": "615",
        "action_path": "/serviceProcess.do?option=initPending4validateVinNo",
        "action_method": "initPending4validateVinNo",
        "action_class": "action.serviceAction"
    },
    {
        "tile_id": 9,
        "tile_name": "Re-Open Job Card",
        "functionality_code": "617",
        "action_path": "/serviceProcess.do?option=initReOpenJobCards",
        "action_method": "initReOpenJobCards",
        "action_class": "action.serviceAction"
    },
    {
        "tile_id": 10,
        "tile_name": "Generate Invoice",
        "functionality_code": "618",
        "action_path": "/serviceProcess.do?option=initCreateInvoiceList",
        "action_method": "initCreateInvoiceList",
        "action_class": "action.serviceAction"
    },
    {
        "tile_id": 11,
        "tile_name": "Tractor History",
        "functionality_code": "620",
        "action_path": "/serviceProcess.do?option=initSearchHistory",
        "action_method": "initSearchHistory",
        "action_class": "action.serviceAction"
    },
    {
        "tile_id": 12,
        "tile_name": "Service Reminder",
        "functionality_code": "621",
        "action_path": "/serviceProcess.do?option=initsercviceReminder",
        "action_method": "initsercviceReminder",
        "action_class": "action.serviceAction"
    },
    {
        "tile_id": 13,
        "tile_name": "Service Done Lapse Report",
        "functionality_code": "622",
        "action_path": "/serviceProcess.do?option=initServiceDoneLapse",
        "action_method": "initServiceDoneLapse",
        "action_class": "action.serviceAction"
    }
]

# API Methods mapping for serviceAction
SERVICE_APIS = {
    "init_viewallJobcarddetails": {
        "request_params": ["vinid", "flag", "range", "jobCardStatus", "fromdate", "todate", "Srch", "dealercode", "user_id"],
        "response": "ArrayList<serviceForm> jobcardDetails",
        "dao_methods": ["getJobCardDetails"],
        "db_tables": ["Jobcarddetails", "Dealervslocationcode", "UmUserDealerMatrix", "SpInventSaleMaster"],
        "stored_procedures": []
    },
    "initCloseJobCards": {
        "request_params": ["user_id", "dealercode", "jobCardStatus"],
        "response": "ArrayList<serviceForm> jobcardDetails",
        "dao_methods": ["getJobCardDetails"],
        "db_tables": ["Jobcarddetails"],
        "stored_procedures": []
    },
    "initReOpenJobCards": {
        "request_params": ["user_id", "vinid", "dealercode"],
        "response": "ArrayList<serviceForm> jobcardDetails",
        "dao_methods": ["getJobCardOpenList"],
        "db_tables": ["Jobcarddetails", "SpInventSaleMaster"],
        "stored_procedures": []
    },
    "initPending4validateVinNo": {
        "request_params": ["user_id", "dealercode", "userFunctionalities"],
        "response": "ArrayList<serviceForm> vinList4Validate",
        "dao_methods": ["getVinNo4validate"],
        "db_tables": ["Jobcarddetails", "Vehicledetails"],
        "stored_procedures": []
    },
    "validateVinNoNupdate": {
        "request_params": ["dealerCode", "vinNo", "jobCardNo", "oldvinNo", "user_id"],
        "response": "String result (success/failure)",
        "dao_methods": ["validateVinNoNupdate"],
        "db_tables": ["Jobcarddetails", "Vehicledetails"],
        "stored_procedures": []
    },
    "initCreateInvoiceList": {
        "request_params": ["user_id", "dealercode"],
        "response": "ArrayList<serviceForm> jobcardDetails",
        "dao_methods": ["getJobCardOpenList"],
        "db_tables": ["Jobcarddetails"],
        "stored_procedures": []
    },
    "generateInvoice": {
        "request_params": ["user_id", "dealercode", "jobCardNo", "vinNo", "jobType", "partNo[]", "partDesc[]", "quantityS[]", "unitPrice[]", "invoicedate", "totalEstimate", "creditValue", "gstNo"],
        "response": "String result (SUCCESS@@Invoice Generated@@invoiceNo / FAILURE / EXIST)",
        "dao_methods": ["generateInvoice"],
        "db_tables": ["SpInventSaleMaster", "SpInventSaleDetails", "Jobcarddetails", "Vehicledetails", "SWVehicleServiceSchedule", "Actualslabourcharges", "Actualsothercharges", "CatPart"],
        "stored_procedures": ["SP_getVehicleCustomerID", "SP_updateTaxInvoice", "SP_BajajExtWtyTaxInvoiceUpdate"]
    },
    "initSearchHistory": {
        "request_params": ["user_id", "dealercode", "vinNo", "userFunctionalities"],
        "response": "ArrayList<serviceForm> vehicle_HistoryList",
        "dao_methods": ["getJobCardHistory"],
        "db_tables": ["Jobcarddetails", "Vehicledetails"],
        "stored_procedures": []
    },
    "initsercviceReminder": {
        "request_params": ["user_id", "dealercode", "fromdate", "todate", "range", "userFunctionalities"],
        "response": "ArrayList<serviceForm> sercviceReminderList",
        "dao_methods": ["getsercviceReminder", "getCommonLabelValues"],
        "db_tables": ["SWVehicleServiceSchedule", "Vehicledetails", "Jobtypemaster"],
        "stored_procedures": []
    },
    "initServiceDoneLapse": {
        "request_params": ["user_id", "dealercode", "fromdate", "todate", "range", "userFunctionalities"],
        "response": "ArrayList<serviceForm> sercviceReminderList",
        "dao_methods": ["getServiceDoneLapse", "getCommonLabelValues"],
        "db_tables": ["SWVehicleServiceSchedule", "Vehicledetails", "Jobtypemaster"],
        "stored_procedures": []
    },
    "viewJobcard": {
        "request_params": ["vinNo", "jobCardNo", "status", "dealerCode", "flag", "pagename", "createInvoice"],
        "response": "serviceForm jobcardDetails",
        "dao_methods": ["viewJobcard", "getCustomerDetailById", "getHesConstantValue"],
        "db_tables": ["Jobcarddetails", "Vehicledetails", "SpInventSaleMaster"],
        "stored_procedures": []
    },
    "getVinDetails": {
        "request_params": ["vinNo", "dealerCode"],
        "response": "String (EngineNo@@DeliveryDate@@RegNo@@SerBookNo@@KeyIdentification@@OwnerDriven@@ModelFamily@@ModelCode@@ModelFamilyDesc@@ModelDesc@@DealerCode@@NextService@@Vinid@@VinDetailsType)",
        "dao_methods": ["getVinDetails"],
        "db_tables": ["Vehicledetails"],
        "stored_procedures": []
    },
    "getvinNumberAjax": {
        "request_params": ["vinNo", "jctype", "user_id", "dealerCode"],
        "response": "String JSON/CSV list of VIN numbers",
        "dao_methods": ["getVinNoList"],
        "db_tables": ["Vehicledetails", "Jobcarddetails"],
        "stored_procedures": []
    },
    "getPartNumberAjax": {
        "request_params": ["partno", "comptype"],
        "response": "String JSON/CSV list of part numbers",
        "dao_methods": ["getComponentList"],
        "db_tables": ["Partmaster", "CatPart"],
        "stored_procedures": []
    },
    "getPartPriceBypartNo": {
        "request_params": ["partno", "priceListCode"],
        "response": "String price value",
        "dao_methods": ["getPriceByPartNo"],
        "db_tables": ["SpPriceMaster", "CatPart"],
        "stored_procedures": []
    },
    "initviewInvoiceList": {
        "request_params": ["user_id", "dealercode", "fromdate", "todate", "range", "creditAmount", "invoiceno", "userFunctionalities"],
        "response": "ArrayList<serviceForm> invoiceList4print",
        "dao_methods": ["getInvNoList4print"],
        "db_tables": ["SpInventSaleMaster", "SpInventSaleDetails"],
        "stored_procedures": []
    },
    "printInvoice": {
        "request_params": ["invoiceNo", "refNo", "type", "dealerCode"],
        "response": "serviceForm jobcardDetails",
        "dao_methods": ["printInvoiceData"],
        "db_tables": ["SpInventSaleMaster", "SpInventSaleDetails", "Jobcarddetails"],
        "stored_procedures": []
    },
    "invoiceTransctionList": {
        "request_params": ["user_id", "dealercode", "fromdate", "todate", "invoiceno", "refNo", "userFunctionalities"],
        "response": "ArrayList<serviceForm> invoiceList4print",
        "dao_methods": ["invTransListByRefNo", "counterSaleReturnList", "invOtherDealerList", "invoiceGrnList", "invExportDealerList"],
        "db_tables": ["SpInventSaleMaster", "SpOrderInvGrn"],
        "stored_procedures": []
    },
    "getFollowup": {
        "request_params": ["scheduleID", "dfleq"],
        "response": "LinkedHashSet<LabelValueBean> serviceTypeIdList, ArrayList followUpHistoryList",
        "dao_methods": ["getCommonLabelValueHiber", "followUpHistoryList"],
        "db_tables": ["Servicetypemaster", "SWVehicleServiceFollowup"],
        "stored_procedures": []
    },
    "updateFollowUp": {
        "request_params": ["user_id", "serviceForm fields"],
        "response": "String result (success/failure)",
        "dao_methods": ["addFollowUp"],
        "db_tables": ["SWVehicleServiceFollowup"],
        "stored_procedures": []
    },
    "setJobCardNoStatus": {
        "request_params": ["jobCardNo", "status", "user_id", "remarks"],
        "response": "String result (success/FAILURE/EXIST)",
        "dao_methods": ["setJobCardNoStatus"],
        "db_tables": ["Jobcarddetails"],
        "stored_procedures": ["SP_UpdateJobCardDailyConsumption"]
    },
    "approveJobCard": {
        "request_params": ["serviceForm", "user_id"],
        "response": "String result (success/failure)",
        "dao_methods": ["approveJobCard"],
        "db_tables": ["Jobcarddetails"],
        "stored_procedures": []
    }
}

# User-defined Classes and Methods
USER_CLASSES = {
    "action.serviceAction": [
        "getPartNumberAjax", "getPartPriceBypartNo", "getPartCheck", "getPartDescAjax",
        "getPartPriceBypartDesc", "init_viewallJobcarddetails", "getVimNumberDetails",
        "getvinNumberAjax", "getServiceHrsAjax", "initPending4validateVinNo",
        "validateVinNoNupdate", "initCloseJobCards", "initReOpenJobCards",
        "initSearchHistory", "viewJobcard", "setJobCardNoStatus", "initCreateInvoiceList",
        "approveJobCard", "generateInvoice", "initviewInvoiceList", "printInvoice",
        "printTrasDetails", "invoiceTransctionList", "initsercviceReminder",
        "initServiceDoneLapse", "printGrn", "getFollowup", "updateFollowUp",
        "printTaxInvoice", "initAddExtWarranty", "addExtWarranty", "getPremiumAmt",
        "isVehicleExist", "isChassisExist", "viewExtendedWarranty", "getExtWarrPopupView",
        "editExtWarranty", "updateExtWarranty", "initAddItlExtWarranty", "addItlExtWarranty",
        "viewItlExtendedWarranty", "fetchDistricts", "fetchCities", "fetchTehsils",
        "isChassisExistInItlExtWty", "getItlExtWarrPopupView", "editItlExtWarranty",
        "updateItlExtWarranty", "printGstTaxInvoice", "isVehicleExistForItlExtWty"
    ],
    "dao.serviceDAO": [
        "getJobCardDetails", "getJobCardOpenList", "getJobCardHistory", "getVinDetails",
        "getVinNoList", "getComponentList", "getPriceByPartNo", "getPartNoCheck",
        "getPart_in_db", "generateInvoice", "getInvNoList4print", "printInvoiceData",
        "invTransListByRefNo", "counterSaleReturnList", "invOtherDealerList",
        "invoiceGrnList", "invExportDealerList", "getsercviceReminder", "getServiceDoneLapse",
        "getCommonLabelValues", "getCommonLabelValueHiber", "followUpHistoryList",
        "addFollowUp", "setJobCardNoStatus", "approveJobCard", "validateVinNoNupdate",
        "getVinNo4validate", "getServiceHrsById", "getCustomerDetailById",
        "getHesConstantValue", "getInvoiceGrn", "getInvoiceGrnDetailsList",
        "printConSaleReturn", "printExportDetails", "printInvOtherDealer",
        "saveExtWarranty", "getPremiumAmt", "isVehicleExist", "isChassisAlreadyExist",
        "getViewExpWarrDetails", "getExpWarrDetailsExport", "getExtWarrPopupView",
        "getPolicyType", "getVehicleDetails", "updateExtWarranty", "saveItlExtWarranty",
        "getViewExpItlExtWarrDetails", "getItlExtWarrPopupView", "updateItlExtWarranty",
        "isChassisExistInItlExtWty", "isVehicleExistForItlExtWty", "getDistricts",
        "getCities", "getTehsils", "getStates", "getExtendedWarrantyDetails",
        "getMechanicCode", "getExtWarrStatus"
    ],
    "com.common.MethodUtility": [
        "getDealersDetailsUnderUser", "check_in_Db_HQL", "getNumber", "getNumberEW",
        "get12DigitInvoiceNoForSpareInvoice", "inventoryLedgerEntry", "checkInDb",
        "getStateIdByDealer"
    ],
    "com.EAMG.common.commonUtilMethods": [
        "decodeBase64", "encodeToBase64"
    ],
    "beans.serviceForm": [
        "All getter and setter methods for form fields"
    ]
}

# Database Tables
DB_TABLES = [
    "Jobcarddetails", "Vehicledetails", "SpInventSaleMaster", "SpInventSaleDetails",
    "Dealervslocationcode", "UmUserDealerMatrix", "SWVehicleServiceSchedule",
    "SWVehicleServiceFollowup", "Jobcardstatusmaster", "Jobtypemaster",
    "Servicetypemaster", "Partmaster", "CatPart", "SpPriceMaster",
    "Actualslabourcharges", "Actualsothercharges", "Estimatelabourcharges",
    "Estimateothercharges", "JobcardSparesCharges", "JobcardSparesActualcharges",
    "Jobcomplaintdetails", "Jobchecklist", "Complaintcodemaster",
    "ApplicationMaster", "Baymaster", "Billmaster", "Causalcodemaster",
    "Consequencemaster", "Joblocationmaster", "Nextservicemaster",
    "Ownerdrivenmaster", "Otherjobworkmaster", "MSWDmechanicmaster",
    "SpOrderInvGrn", "SpOrderInvGrnDetails", "EWMLoaderDetail",
    "ITLEWMLoaderDetail", "ItlContentMaster", "ItlSubContentMaster",
    "TMSDelayReasonMaster", "SpCurrentInventoryView"
]

# Stored Procedures
STORED_PROCEDURES = [
    "SP_getVehicleCustomerID", "SP_updateTaxInvoice", "SP_BajajExtWtyTaxInvoiceUpdate",
    "SP_UpdateJobCardDailyConsumption", "SP_showTaxInvoice", "SP_showTaxValuesInInvoice",
    "SP_JobCardComplaintMandatoryCheck", "SP_CheckEstimateMandatory"
]

def create_excel_file():
    """Create comprehensive Excel file with service module analysis"""
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    title_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Sheet 1: Service Tiles Overview
    ws1 = wb.create_sheet("1. Service Tiles Overview")
    headers1 = ["Tile ID", "Tile Name", "Functionality Code", "Action Path", "Action Method", "Action Class"]
    ws1.append(headers1)
    
    for header_cell in ws1[1]:
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border
    
    for tile in SERVICE_TILES:
        row = [
            tile["tile_id"],
            tile["tile_name"],
            tile["functionality_code"],
            tile["action_path"],
            tile["action_method"],
            tile["action_class"]
        ]
        ws1.append(row)
        for cell in ws1[ws1.max_row]:
            cell.border = border
            cell.alignment = wrap_alignment
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 70
    ws1.column_dimensions['E'].width = 30
    ws1.column_dimensions['F'].width = 30
    
    # Sheet 2: APIs Called per Tile
    ws2 = wb.create_sheet("2. APIs Called per Tile")
    headers2 = ["Tile ID", "Tile Name", "API Method", "Request Parameters", "Response", "DAO Methods Called"]
    ws2.append(headers2)
    
    for header_cell in ws2[1]:
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border
    
    row_num = 2
    for tile in SERVICE_TILES:
        method = tile["action_method"]
        if method in SERVICE_APIS:
            api_info = SERVICE_APIS[method]
            row = [
                tile["tile_id"],
                tile["tile_name"],
                method,
                ", ".join(api_info["request_params"]),
                api_info["response"],
                ", ".join(api_info["dao_methods"])
            ]
            ws2.append(row)
            for cell in ws2[ws2.max_row]:
                cell.border = border
                cell.alignment = wrap_alignment
            row_num += 1
    
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 50
    ws2.column_dimensions['E'].width = 50
    ws2.column_dimensions['F'].width = 50
    
    # Sheet 3: All APIs Summary
    ws3 = wb.create_sheet("3. All APIs Summary")
    headers3 = ["API Method", "Request Parameters", "Response", "DAO Methods", "DB Tables", "Stored Procedures"]
    ws3.append(headers3)
    
    for header_cell in ws3[1]:
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border
    
    for method, api_info in SERVICE_APIS.items():
        row = [
            method,
            ", ".join(api_info["request_params"]),
            api_info["response"],
            ", ".join(api_info["dao_methods"]),
            ", ".join(api_info["db_tables"]),
            ", ".join(api_info["stored_procedures"]) if api_info["stored_procedures"] else "None"
        ]
        ws3.append(row)
        for cell in ws3[ws3.max_row]:
            cell.border = border
            cell.alignment = wrap_alignment
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 50
    ws3.column_dimensions['D'].width = 50
    ws3.column_dimensions['E'].width = 50
    ws3.column_dimensions['F'].width = 50
    
    # Sheet 4: User-Defined Classes and Methods
    ws4 = wb.create_sheet("4. User Classes & Methods")
    headers4 = ["Class Name", "Methods"]
    ws4.append(headers4)
    
    for header_cell in ws4[1]:
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border
    
    for class_name, methods in USER_CLASSES.items():
        row = [
            class_name,
            ", ".join(methods)
        ]
        ws4.append(row)
        for cell in ws4[ws4.max_row]:
            cell.border = border
            cell.alignment = wrap_alignment
    
    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 100
    
    # Sheet 5: Database Tables
    ws5 = wb.create_sheet("5. Database Tables")
    headers5 = ["Table Name"]
    ws5.append(headers5)
    
    for header_cell in ws5[1]:
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border
    
    for table in sorted(DB_TABLES):
        row = [table]
        ws5.append(row)
        for cell in ws5[ws5.max_row]:
            cell.border = border
            cell.alignment = wrap_alignment
    
    ws5.column_dimensions['A'].width = 40
    
    # Sheet 6: Stored Procedures
    ws6 = wb.create_sheet("6. Stored Procedures")
    headers6 = ["Stored Procedure Name"]
    ws6.append(headers6)
    
    for header_cell in ws6[1]:
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border
    
    for sp in sorted(STORED_PROCEDURES):
        row = [sp]
        ws6.append(row)
        for cell in ws6[ws6.max_row]:
            cell.border = border
            cell.alignment = wrap_alignment
    
    ws6.column_dimensions['A'].width = 50
    
    # Sheet 7: Tile to API Mapping
    ws7 = wb.create_sheet("7. Tile-API-DB Mapping")
    headers7 = ["Tile ID", "Tile Name", "Primary API", "Request Params", "Response", "DB Tables Used", "Stored Procedures"]
    ws7.append(headers7)
    
    for header_cell in ws7[1]:
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border
    
    for tile in SERVICE_TILES:
        method = tile["action_method"]
        if method in SERVICE_APIS:
            api_info = SERVICE_APIS[method]
            row = [
                tile["tile_id"],
                tile["tile_name"],
                method,
                ", ".join(api_info["request_params"][:5]) + ("..." if len(api_info["request_params"]) > 5 else ""),
                api_info["response"][:50] + ("..." if len(api_info["response"]) > 50 else ""),
                ", ".join(api_info["db_tables"][:5]) + ("..." if len(api_info["db_tables"]) > 5 else ""),
                ", ".join(api_info["stored_procedures"]) if api_info["stored_procedures"] else "None"
            ]
            ws7.append(row)
            for cell in ws7[ws7.max_row]:
                cell.border = border
                cell.alignment = wrap_alignment
    
    ws7.column_dimensions['A'].width = 10
    ws7.column_dimensions['B'].width = 30
    ws7.column_dimensions['C'].width = 30
    ws7.column_dimensions['D'].width = 40
    ws7.column_dimensions['E'].width = 40
    ws7.column_dimensions['F'].width = 50
    ws7.column_dimensions['G'].width = 50
    
    # Save the workbook
    filename = "Service_Module_Analysis.xlsx"
    wb.save(filename)
    print(f"Excel file '{filename}' created successfully!")
    print(f"\nSummary:")
    print(f"- Total Tiles: {len(SERVICE_TILES)}")
    print(f"- Total APIs: {len(SERVICE_APIS)}")
    print(f"- Total User Classes: {len(USER_CLASSES)}")
    print(f"- Total DB Tables: {len(DB_TABLES)}")
    print(f"- Total Stored Procedures: {len(STORED_PROCEDURES)}")

if __name__ == "__main__":
    try:
        create_excel_file()
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()

