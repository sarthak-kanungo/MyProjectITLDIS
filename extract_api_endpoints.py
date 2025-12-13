import os
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def extract_api_endpoints(backend_path):
    """Extract all API endpoints from Java controller files"""
    endpoints = []
    
    # Pattern to find controller files
    controller_files = []
    for root, dirs, files in os.walk(backend_path):
        for file in files:
            if file.endswith('Controller.java'):
                controller_files.append(os.path.join(root, file))
    
    print(f"Found {len(controller_files)} controller files")
    
    for file_path in controller_files:
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Extract base path from @RequestMapping
            base_path_match = re.search(r'@RequestMapping\s*\([^)]*value\s*=\s*["\']([^"\']+)["\']', content, re.MULTILINE)
            if not base_path_match:
                # Try without value attribute
                base_path_match = re.search(r'@RequestMapping\s*\(["\']([^"\']+)["\']', content, re.MULTILINE)
            
            base_path = base_path_match.group(1) if base_path_match else ""
            
            # Extract class name
            class_match = re.search(r'public\s+class\s+(\w+Controller)', content)
            class_name = class_match.group(1) if class_match else "Unknown"
            
            # Find all mapping annotations with their methods
            # Pattern: @MappingAnnotation(...) followed by public method
            mapping_pattern = r'@(Get|Post|Put|Delete|Patch)Mapping\s*\([^)]*\)'
            
            # Find all occurrences
            for match in re.finditer(mapping_pattern, content, re.MULTILINE | re.DOTALL):
                mapping_annotation = match.group(0)
                mapping_type = match.group(1).upper()
                
                # Extract path from annotation
                endpoint_path = ""
                # Try value = "path"
                path_match = re.search(r'value\s*=\s*["\']([^"\']+)["\']', mapping_annotation)
                if not path_match:
                    # Try just "path"
                    path_match = re.search(r'["\']([^"\']+)["\']', mapping_annotation)
                if path_match:
                    endpoint_path = path_match.group(1)
                
                # Find the method that follows this annotation
                # Look for public method after this annotation
                method_start = match.end()
                method_content = content[method_start:method_start+1000]
                
                # Find method signature
                method_sig_match = re.search(r'public\s+([^{]+)\{', method_content, re.MULTILINE | re.DOTALL)
                if method_sig_match:
                    method_signature = method_sig_match.group(1).strip()
                    
                    # Extract request and response info
                    request_info = extract_request_info(method_signature)
                    response_info = extract_response_info(method_signature)
                    
                    # Build full endpoint
                    full_endpoint = base_path
                    if endpoint_path:
                        if endpoint_path.startswith('/'):
                            full_endpoint += endpoint_path
                        else:
                            full_endpoint += '/' + endpoint_path
                    elif not full_endpoint.endswith('/'):
                        full_endpoint += '/'
                    
                    endpoints.append({
                        'controller': class_name,
                        'http_method': mapping_type,
                        'endpoint': full_endpoint,
                        'request': request_info,
                        'response': response_info,
                        'file': os.path.basename(file_path)
                    })
            
            # Also handle @RequestMapping with method attribute
            request_mapping_pattern = r'@RequestMapping\s*\([^)]*method\s*=\s*RequestMethod\.(\w+)'
            for match in re.finditer(request_mapping_pattern, content, re.MULTILINE | re.DOTALL):
                mapping_annotation = content[match.start():match.end()+200]  # Get more context
                mapping_type = match.group(1).upper()
                
                # Extract path
                endpoint_path = ""
                path_match = re.search(r'value\s*=\s*["\']([^"\']+)["\']', mapping_annotation)
                if not path_match:
                    path_match = re.search(r'["\']([^"\']+)["\']', mapping_annotation)
                if path_match:
                    endpoint_path = path_match.group(1)
                
                # Find method
                method_start = match.end()
                method_content = content[method_start:method_start+1000]
                method_sig_match = re.search(r'public\s+([^{]+)\{', method_content, re.MULTILINE | re.DOTALL)
                
                if method_sig_match:
                    method_signature = method_sig_match.group(1).strip()
                    request_info = extract_request_info(method_signature)
                    response_info = extract_response_info(method_signature)
                    
                    full_endpoint = base_path
                    if endpoint_path:
                        if endpoint_path.startswith('/'):
                            full_endpoint += endpoint_path
                        else:
                            full_endpoint += '/' + endpoint_path
                    elif not full_endpoint.endswith('/'):
                        full_endpoint += '/'
                    
                    endpoints.append({
                        'controller': class_name,
                        'http_method': mapping_type,
                        'endpoint': full_endpoint,
                        'request': request_info,
                        'response': response_info,
                        'file': os.path.basename(file_path)
                    })
        
        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")
            continue
    
    return endpoints

def extract_request_info(method_signature):
    """Extract request information from method signature"""
    if not method_signature:
        return "None"
    
    request_info = []
    
    # Extract @RequestBody parameters
    request_body_matches = re.findall(r'@RequestBody\s+(?:\w+\s+)?(\w+(?:<[^>]+>)?)\s+(\w+)', method_signature)
    for param_type, param_name in request_body_matches:
        request_info.append(f"@RequestBody {param_name}: {param_type}")
    
    # Extract @RequestParam parameters
    request_params = re.findall(r'@RequestParam(?:\([^)]*\))?\s+(?:\w+\s+)?(\w+(?:<[^>]+>)?)\s+(\w+)', method_signature)
    for param_type, param_name in request_params:
        request_info.append(f"@RequestParam {param_name}: {param_type}")
    
    # Extract @PathVariable parameters
    path_vars = re.findall(r'@PathVariable(?:\([^)]*\))?\s+(?:\w+\s+)?(\w+(?:<[^>]+>)?)\s+(\w+)', method_signature)
    for var_type, var_name in path_vars:
        request_info.append(f"@PathVariable {var_name}: {var_type}")
    
    # Extract @RequestPart (for file uploads)
    request_parts = re.findall(r'@RequestPart(?:\([^)]*\))?\s+(?:\w+\s+)?(\w+(?:<[^>]+>)?)\s+(\w+)', method_signature)
    for part_type, part_name in request_parts:
        request_info.append(f"@RequestPart {part_name}: {part_type}")
    
    # Extract MultipartFile
    multipart_matches = re.findall(r'MultipartFile\s+(\w+)', method_signature)
    for file_name in multipart_matches:
        request_info.append(f"MultipartFile: {file_name}")
    
    # Extract HttpServletRequest, HttpServletResponse
    if 'HttpServletRequest' in method_signature:
        request_info.append("HttpServletRequest")
    if 'HttpServletResponse' in method_signature:
        request_info.append("HttpServletResponse")
    
    # Extract simple parameters (fallback)
    if not request_info:
        # Try to extract parameters from method signature
        params_section = re.search(r'\(([^)]+)\)', method_signature)
        if params_section:
            params = params_section.group(1)
            # Split by comma and extract type and name
            param_list = re.findall(r'(\w+(?:<[^>]+>)?)\s+(\w+)', params)
            for param_type, param_name in param_list:
                if param_type not in ['public', 'ResponseEntity', 'ApiResponse', 'void', 'final']:
                    request_info.append(f"{param_name}: {param_type}")
    
    return ", ".join(request_info) if request_info else "None"

def extract_response_info(method_signature):
    """Extract response information from method signature"""
    if not method_signature:
        return "Unknown"
    
    # Extract return type - ResponseEntity<Type>
    return_match = re.search(r'ResponseEntity\s*<([^>]+)>', method_signature)
    if return_match:
        return return_match.group(1).strip()
    
    # Extract simple return type
    return_match = re.search(r'public\s+(\w+(?:<[^>]+>)?)', method_signature)
    if return_match:
        return_type = return_match.group(1).strip()
        if return_type not in ['void', 'class']:
            return return_type
    
    return "ResponseEntity<?>"

def create_excel(endpoints, output_file):
    """Create Excel file with API endpoints"""
    wb = Workbook()
    ws = wb.active
    ws.title = "API Endpoints"
    
    # Headers
    headers = ['S.No', 'HTTP Method', 'API Endpoint', 'Request', 'Response', 'Controller']
    
    # Style for header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for idx, endpoint in enumerate(endpoints, 2):
        ws.cell(row=idx, column=1, value=idx-1)  # S.No
        ws.cell(row=idx, column=2, value=endpoint['http_method'])  # HTTP Method
        ws.cell(row=idx, column=3, value=endpoint['endpoint'])  # API Endpoint
        ws.cell(row=idx, column=4, value=endpoint['request'])  # Request
        ws.cell(row=idx, column=5, value=endpoint['response'])  # Response
        ws.cell(row=idx, column=6, value=endpoint['controller'])  # Controller
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel file created: {output_file}")

if __name__ == "__main__":
    backend_path = "ITLDIS-BACKEND/src/main/java"
    
    if not os.path.exists(backend_path):
        print(f"Backend path not found: {backend_path}")
        exit(1)
    
    print("Extracting API endpoints...")
    endpoints = extract_api_endpoints(backend_path)
    
    print(f"\nTotal endpoints found: {len(endpoints)}")
    
    # Remove duplicates based on endpoint and method
    seen = set()
    unique_endpoints = []
    for ep in endpoints:
        key = (ep['endpoint'], ep['http_method'])
        if key not in seen:
            seen.add(key)
            unique_endpoints.append(ep)
    
    print(f"Unique endpoints: {len(unique_endpoints)}")
    
    # Create Excel file
    output_file = "ITLDIS_API_Endpoints.xlsx"
    create_excel(unique_endpoints, output_file)
    
    print(f"\nSummary:")
    print(f"- Total API Endpoints: {len(unique_endpoints)}")
    print(f"- Excel file saved as: {output_file}")
